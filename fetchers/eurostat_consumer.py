"""
Fetches consumer inflation expectations (balance scores) from Eurostat
Consumer Survey zip file.

Source: https://ec.europa.eu/economy_finance/db_indicators/surveys/documents/
         series/nace2_ecfin_{YYYYMM}/consumer_inflation_nace2.zip
File:   consumer_subsectors_nsa_q6_nace2.xlsx  →  sheet 'TOT'
Column: CONS.{CC}.TOT.6.B.M  (balance = % expecting higher − % expecting lower)

The zip (~28MB) is downloaded ONCE and all countries parsed together;
results are cached per country for 12 hours.

Series id = 2-letter country code: EA, PL, HU, CZ, RO
"""

import io
import logging
import zipfile
from datetime import date

import requests
import cache
from config import DEFAULT_START_DATE

logger = logging.getLogger(__name__)

import threading
_download_lock = threading.Lock()   # prevent concurrent 28MB downloads

# URL uses YYYYMM of the latest release — update the suffix when Eurostat republishes
ZIP_URL = (
    "https://ec.europa.eu/economy_finance/db_indicators/surveys/documents/"
    "series/nace2_ecfin_2606/consumer_inflation_nace2.zip"
)
XLSX_NAME = "consumer_subsectors_nsa_q6_nace2.xlsx"
SHEET     = "TOT"
DATE_COL  = 1   # Column A: monthly dates

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; cb-api/1.0)"}

# All-countries cache key — shared across country fetches to avoid re-downloading
ALL_CACHE_KEY = f"eurostat_consumer_all_{DEFAULT_START_DATE[:7]}"


def _download_all(start_date: str) -> dict:
    """
    Download the zip, parse the xlsx, and return a dict:
    { "EA": [{date, value}, ...], "PL": [...], ... }
    Cached for 12 hours under ALL_CACHE_KEY.
    """
    cached = cache.get(ALL_CACHE_KEY)
    if cached is not None:
        return cached

    with _download_lock:
        # Re-check inside the lock in case another thread just finished
        cached = cache.get(ALL_CACHE_KEY)
        if cached is not None:
            return cached

        logger.info("Downloading Eurostat consumer survey zip (~28MB)…")
    resp = requests.get(ZIP_URL, headers=HEADERS, timeout=120)
    resp.raise_for_status()

    import openpyxl
    with zipfile.ZipFile(io.BytesIO(resp.content)) as z:
        with z.open(XLSX_NAME) as f:
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)

    ws = wb[SHEET]
    # Row 1 = column headers;  row 2+ = data
    row1 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # Find column indices for each target country code
    country_cols = {}
    for cc in ("EA", "PL", "HU", "CZ", "RO"):
        key = f"CONS.{cc}.TOT.6.B.M"
        if key in row1:
            country_cols[cc] = row1.index(key) + 1   # 1-based

    cutoff = start_date[:7]
    result = {cc: [] for cc in country_cols}

    for row_idx in range(2, ws.max_row + 1):
        raw_date = ws.cell(row_idx, DATE_COL).value
        if not raw_date:
            continue
        # Dates are datetime objects from openpyxl
        if hasattr(raw_date, "strftime"):
            iso = raw_date.strftime("%Y-%m-%d")
        else:
            continue
        if iso[:7] < cutoff:
            continue
        for cc, col_idx in country_cols.items():
            val = ws.cell(row_idx, col_idx).value
            if val is not None:
                try:
                    result[cc].append({"date": iso, "value": round(float(val), 2)})
                except (ValueError, TypeError):
                    pass

    for cc in result:
        result[cc].sort(key=lambda r: r["date"])

    cache.set(ALL_CACHE_KEY, result)
    logger.info("Eurostat consumer survey: %s countries, latest %s",
                list(result.keys()), next(iter(result.values()), [{}])[-1:])
    return result


def fetch(country_code: str, start_date: str = DEFAULT_START_DATE) -> list[dict]:
    """
    Return [{date, value}] for the given country code (EA, PL, HU, CZ, RO).
    Balance score: positive = more people expect higher inflation than lower.
    """
    cc = country_code.upper()
    cache_key = f"eurostat_consumer_{cc}_{start_date[:7]}"
    cached = cache.get(cache_key)
    if cached is not None:
        return cached

    all_data = _download_all(start_date)
    result = all_data.get(cc, [])
    logger.debug("Eurostat consumer %s: %d monthly observations", cc, len(result))
    cache.set(cache_key, result)
    return result
