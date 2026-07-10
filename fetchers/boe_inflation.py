"""
Fetches BoE Inflation Attitudes Survey — 1-year median inflation expectation.
Source: https://www.bankofengland.co.uk/-/media/boe/files/inflation-attitudes-survey/long-run.xlsx

Row 4  = survey dates (datetime objects → YYYY-MM-DD)
Row 37 = median 1-year inflation expectation (%)
Frequency: quarterly.
"""

import io
import logging
from datetime import date

import requests
import cache
from config import DEFAULT_START_DATE

logger = logging.getLogger(__name__)

URL     = "https://www.bankofengland.co.uk/-/media/boe/files/inflation-attitudes-survey/long-run.xlsx"
HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; cb-api/1.0)"}


def fetch(series_id: str = "long-run-median", start_date: str = DEFAULT_START_DATE) -> list[dict]:
    cache_key = f"boe_inflation_{start_date[:7]}"
    cached = cache.get(cache_key)
    if cached is not None:
        return cached

    logger.info("Fetching BoE Inflation Attitudes Survey xlsx")
    resp = requests.get(URL, headers=HEADERS, timeout=30)
    resp.raise_for_status()

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
    ws = wb.active

    cutoff = start_date[:10]
    results = []

    for col in range(2, ws.max_column + 1):
        raw_date   = ws.cell(4,  col).value
        raw_median = ws.cell(37, col).value

        if not raw_date or raw_median is None:
            continue

        # Convert Excel date (datetime) to ISO string
        if hasattr(raw_date, "strftime"):
            iso = raw_date.strftime("%Y-%m-%d")
        else:
            continue

        if iso < cutoff:
            continue

        try:
            results.append({"date": iso, "value": round(float(raw_median), 4)})
        except (ValueError, TypeError):
            pass

    results.sort(key=lambda r: r["date"])
    logger.debug("BoE inflation expectations: %d quarterly observations", len(results))
    cache.set(cache_key, results)
    return results
