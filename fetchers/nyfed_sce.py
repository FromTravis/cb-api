"""
Fetches NY Fed Survey of Consumer Expectations (SCE) inflation expectations.
Source: https://www.newyorkfed.org/medialibrary/interactives/sce/sce/downloads/data/frbny-sce-data.xlsx
Sheet:  'Inflation expectations'
Column: 'Median one-year ahead expected inflation rate' (column B)
Date:   Column A as integer YYYYMM (e.g. 202306 → 2023-06-01)
Frequency: monthly.
"""

import io
import logging

import requests
import cache
from config import DEFAULT_START_DATE

logger = logging.getLogger(__name__)

URL      = "https://www.newyorkfed.org/medialibrary/interactives/sce/sce/downloads/data/frbny-sce-data.xlsx"
SHEET    = "Inflation expectations"
HEADERS  = {"User-Agent": "Mozilla/5.0 (compatible; cb-api/1.0)"}
HEADER_ROW = 4   # Row with column names
DATA_START = 5   # First data row


def fetch(series_id: str = "median_1y", start_date: str = DEFAULT_START_DATE) -> list[dict]:
    cache_key = f"nyfed_sce_{series_id}_{start_date[:7]}"
    cached = cache.get(cache_key)
    if cached is not None:
        return cached

    logger.info("Fetching NY Fed SCE xlsx")
    resp = requests.get(URL, headers=HEADERS, timeout=30)
    resp.raise_for_status()

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    ws = wb[SHEET]

    # Find the target column from row 4 headers
    headers_row = [ws.cell(HEADER_ROW, c).value for c in range(1, ws.max_column + 1)]
    target_col = None
    for i, h in enumerate(headers_row):
        if h and "median one-year ahead expected inflation" in str(h).lower():
            target_col = i + 1   # 1-based
            break

    if target_col is None:
        logger.error("Could not find target column in NY Fed SCE sheet")
        return []

    cutoff  = start_date[:7]
    results = []

    for row_idx in range(DATA_START, ws.max_row + 1):
        raw_date = ws.cell(row_idx, 1).value    # YYYYMM integer
        raw_val  = ws.cell(row_idx, target_col).value
        if not raw_date or raw_val is None:
            continue
        try:
            yyyymm = str(int(raw_date))
            iso    = f"{yyyymm[:4]}-{yyyymm[4:6]}-01"
            if iso[:7] < cutoff:
                continue
            results.append({"date": iso, "value": round(float(raw_val), 4)})
        except (ValueError, TypeError):
            pass

    results.sort(key=lambda r: r["date"])
    logger.debug("NY Fed SCE: %d monthly observations", len(results))
    cache.set(cache_key, results)
    return results
